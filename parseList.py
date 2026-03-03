import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import numpy as np
from openpyxl.styles import numbers  # for date formatting
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

output_name = ""

hotel_map = {
    "Adler Resort": "Adler",
    "Hotel Alpen Karawanserai": "AKW",
    "Hotel Sonnberg": "Sonnberg",
}

def select_file():
    global input_path
    input_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if input_path:
        file_label.config(text=os.path.basename(input_path))

def process_file():
    # input_path = filedialog.askopenfilename(
    # filetypes=[("Excel files", "*.xlsx *.xls")]
    # )   
    
    if not input_path:
        messagebox.showerror("Error", "Please select an input file.")
        return
    
    output_name = output_entry.get().strip()
    if not output_name:
        messagebox.showerror("Error", "Please enter an output file name.")
        return
    
    save_folder = filedialog.askdirectory(title="Select folder to save")
    if not save_folder:
        return

    output_path = os.path.join(save_folder, f"{output_name}.xlsx")
    output_path = os.path.normpath(output_path)


    try:
        # Read second sheet, force phone numbers as string
        df = pd.read_excel(
            input_path,
            sheet_name=1,
            dtype={"phoneNumber": str}
        )

        # Create new DataFrame
        new_df = pd.DataFrame()
        new_df["Nummer"] = ""
        new_df["fullName"] = df["fullName"]
        new_df["company"] = df["company"]
        new_df["phoneNumber"] = df["phoneNumber"]
        new_df["hotel"] = df["hotel"]
        new_df["startDate"] = pd.to_datetime(df["startDate"]).dt.date
        new_df["endDate"] = pd.to_datetime(df["endDate"]).dt.date
        new_df["arrivalDate"] = pd.to_datetime(df["arrivalDate"]).dt.date
        new_df["days"] = df["days"]
        new_df["height"] = df["height"]
        new_df["weight"] = df["weight"]
        new_df["shoeSize"] = df["shoeSize"]
        new_df["level"] = df["level"]
        new_df["binding"] = df["binding"]
        new_df["pole"] = df["pole"]
        new_df["adjustment"] = df["adjustment"]
        new_df["shoeType"] = df["shoeType"]
        new_df["schoennummer"] = ""
        new_df["subType"] = df["subType"]
        new_df["skinummer"] = ""
        new_df["helm"] = df["helm"]

        new_df.loc[new_df["shoeType"].isna() | (new_df["shoeType"] == ""), "schoennummer"] = "ES"
        new_df.loc[df["helm"] == True, "helm"] = "Helm"
        new_df["hotel"] = df["hotel"].map(hotel_map)
        new_df.loc[df["height"] == 165, "pole"] = 110

        # Sort by hotel → arrivalDate → fullName
        new_df.sort_values(
            by=["hotel", "startDate", "fullName"],
            inplace=True
        )

        new_df["Nummer"] = new_df.groupby("hotel").cumcount() + 1

        new_df["pole"] = new_df["pole"].astype("string")
        condition = new_df["subType"].isin([
            "Huur snowboard  - gold",
            "Huur snowboard - platinum"
        ])

        new_df.loc[condition, "pole"] = np.select(
            [
                new_df.loc[condition, "shoeSize"].between(35, 39),
                new_df.loc[condition, "shoeSize"].between(40, 43),
                new_df.loc[condition, "shoeSize"] >= 44
            ],
            [
                "Blauw",
                "Rood",
                "Geel"
            ],
            default=""
        )


        # Define output path
        # output_path = os.path.join(
        #     os.path.dirname(input_path),
        #     "numbered_" + os.path.basename(input_path)
        # )

        # Write Excel with auto-width and proper date format
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            new_df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]

            # Auto-adjust all columns
            for col_idx in range(1, len(new_df.columns) + 1):
                max_length = 0
                for row_idx in range(1, len(new_df) + 2):  # +1 for header
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[column_letter].width = max_length + 6

            # Format date columns as DD-MM-YYYY
            date_columns = ["startDate", "endDate", "arrivalDate"]
            for col in date_columns:
                if col in new_df.columns:
                    col_idx = list(new_df.columns).index(col) + 1
                    for row_idx in range(2, len(new_df) + 2):  # skip header
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = "DD-MM-YYYY"


            red_font = Font(color="FF0000")
            bold_font = Font(bold=True)

            gold_board_condition = new_df["subType"] == "Huur snowboard  - gold"
            platinum_ski_condition = new_df["subType"] == "Huur ski - Platinum"
            platinum_board_condition = new_df["subType"] == "Huur snowboard - platinum"
     

            for col_idx in range(1, len(new_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = Font(name="Aptos Narrow", size=14)
            
            for row_idx in range(2, len(new_df) + 2):  # skip header
                row_data = new_df.iloc[row_idx - 2]


                for col_idx in range(1, len(new_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    if gold_board_condition.iloc[row_idx - 2]:
                        cell.font = Font(color="FF0000", name="Aptos Narrow", size=14)

                    elif platinum_ski_condition.iloc[row_idx - 2]:
                        cell.font = Font(bold=True, name="Aptos Narrow", size=14)

                    elif platinum_board_condition.iloc[row_idx - 2]:
                        cell.font = Font(color="FF0000", bold=True, name="Aptos Narrow", size=14)

                    else:
                        cell.font = Font(name="Aptos Narrow", size=14)

        messagebox.showinfo("Success", f"File created:\n{output_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))

    root.destroy()


root = tk.Tk()
root.title("Excel Processor")
root.geometry("400x250")

tk.Label(root, text="Selecteer Fruhstuk Excel bestand:").pack(pady=(10,0))
tk.Button(root, text="Browse", command=select_file).pack()
file_label = tk.Label(root, text="Geen bestand geselecteerd", fg="blue")
file_label.pack(pady=(5,10))

tk.Label(root, text="Voer bestandsnaam in:").pack()
output_entry = tk.Entry(root, width=40)
output_entry.pack(pady=(0,10))

tk.Button(root, text="Genereer", command=process_file, width=20).pack(pady=10)

root.mainloop()