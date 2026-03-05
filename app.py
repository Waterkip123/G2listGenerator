import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import Font, Border, Side, PatternFill

st.set_page_config(
    page_title="Excel Processor", 
    page_icon="g2Logo.ico", 
    layout="centered",     
    initial_sidebar_state="collapsed",
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    })

st.title("Excel Processor")

hotel_map = {
    "Adler Resort": "Adler",
    "Hotel Alpen Karawanserai": "AKW",
    "Hotel Sonnberg": "Sonnberg",
}

uploaded_file = st.file_uploader(
    "Upload Fruhstuk Excel bestand",
    type=["xlsx", "xls", "xlsm"]
)

output_name = st.text_input("Voer bestandsnaam in")

if uploaded_file and output_name:

    try:
        # Read second sheet
        df = pd.read_excel(
            uploaded_file,
            sheet_name=1,
            dtype={"phoneNumber": str}
        )

        new_df = pd.DataFrame()
        new_df["Nummer"] = ""
        new_df["fullName"] = df["fullName"]
        new_df["company"] = df["company"]
        new_df["phoneNumber"] = df["phoneNumber"]
        new_df["hotel"] = df["hotel"]
        new_df["startDate"] = pd.to_datetime(df["startDate"]).dt.date
        new_df["endDate"] = pd.to_datetime(df["endDate"]).dt.date
        new_df["arrivalDate"] = pd.to_datetime(df["arrivalDate"]).dt.date
        new_df["days"] = df["days"]
        new_df["height"] = df["height"]
        new_df["weight"] = df["weight"]
        new_df["shoeSize"] = df["shoeSize"]
        new_df["level"] = df["level"]
        new_df["binding"] = df["binding"]
        new_df["pole"] = df["pole"]
        new_df["adjustment"] = df["adjustment"]
        new_df["shoeType"] = df["shoeType"]
        new_df["schoennummer"] = ""
        new_df["subType"] = df["subType"]
        new_df["skinummer"] = ""
        new_df["helm"] = df["helm"]

        new_df.loc[new_df["shoeType"].isna() | (new_df["shoeType"] == ""), "schoennummer"] = "ES"
        new_df.loc[df["helm"] == True, "helm"] = "Helm"
        new_df["hotel"] = df["hotel"].map(hotel_map)
        new_df.loc[df["height"] == 165, "pole"] = 110

        # Sort
        new_df["hotel_sort"] = new_df["hotel"].str.lower()
        new_df["name_sort"] = new_df["fullName"].str.lower()

        new_df.sort_values(
            by=["hotel_sort", "startDate", "name_sort"],
            inplace=True
        )

        new_df.drop(columns=["hotel_sort", "name_sort"], inplace=True)

        # Nummer per hotel reset
        new_df["Nummer"] = range(1, len(new_df) + 1)
        # Pole logic
        new_df["pole"] = new_df["pole"].astype("string")

        condition = new_df["subType"].isin([
            "Huur snowboard  - gold",
            "Huur snowboard - platinum"
        ])

        new_df.loc[condition, "pole"] = np.select(
            [
                new_df.loc[condition, "shoeSize"].between(35, 39),
                new_df.loc[condition, "shoeSize"].between(40, 43),
                new_df.loc[condition, "shoeSize"] >= 44
            ],
            ["Blauw", "Rood", "Geel"],
            default=""
        )

        # Create Excel in memory
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            new_df.to_excel(writer, index=False, sheet_name="Sheet1")
            worksheet = writer.sheets["Sheet1"]

            # Auto column width
            for col_idx in range(1, len(new_df.columns) + 1):
                max_length = 0
                for row_idx in range(1, len(new_df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[column_letter].width = max_length + 6

            # Date formatting
            date_columns = ["startDate", "endDate", "arrivalDate"]
            for col in date_columns:
                if col in new_df.columns:
                    col_idx = list(new_df.columns).index(col) + 1
                    for row_idx in range(2, len(new_df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = "DD-MM-YYYY"

            # Styling
            gold_board_condition = new_df["subType"] == "Huur snowboard  - gold"
            platinum_ski_condition = new_df["subType"] == "Huur ski - Platinum"
            platinum_board_condition = new_df["subType"] == "Huur snowboard - platinum"

        
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # black
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # light gray

            thin_side = Side(border_style="thin", color="000000")
            thin_border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side
            )


            for col_idx in range(1, len(new_df.columns) + 1):
                worksheet.cell(row=1, column=col_idx).font = Font(name="Aptos Narrow", color="FFFFFF", size=14)
                worksheet.cell(row=1, column=col_idx).fill = header_fill


            for row_idx in range(2, len(new_df) + 2):
                for col_idx in range(1, len(new_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    cell.border = thin_border

                    if row_idx % 2 == 1:
                        cell.fill = gray_fill

                    if gold_board_condition.iloc[row_idx - 2]:
                        cell.font = Font(color="FF0000", name="Aptos Narrow", size=14)

                    elif platinum_ski_condition.iloc[row_idx - 2]:
                        cell.font = Font(bold=True, name="Aptos Narrow", size=14)

                    elif platinum_board_condition.iloc[row_idx - 2]:
                        cell.font = Font(color="FF0000", bold=True, name="Aptos Narrow", size=14)

                    else:
                        cell.font = Font(name="Aptos Narrow", size=14)

                    if pd.isna(new_df["shoeType"].iloc[row_idx - 2]):
                        current = cell.font
                        cell.font = Font(
                            name=current.name,
                            size=current.size,
                            bold=current.bold,
                            italic=True,
                            color=current.color
                        )

        buffer.seek(0)

        st.success("Bestand gereed!")

        st.download_button(
            label="Download Excel bestand",
            data=buffer,
            file_name=f"{output_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(str(e))